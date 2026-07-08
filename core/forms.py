import glob
import os
import openpyxl
import re
from django import forms
from .models import Task, User, Mesaj, Reminder, RaportSupervizor

# --- LISTE AUXILIARE ---
ACTIUNI_LISTA = [
    ('mentenanta', 'Mentenanță'),
    ('instalare_video', 'Instalare Video'),
    ('instalare_efractie', 'Instalare Efractie'),
    ('instalare_control_acces_pietonal', 'Instalare Control-Acces Pietonal'),
    ('instalare_control_acces_auto', 'Instalare Control-Acces Auto'),
    ('schimbare_dvr', 'Schimbare DVR'),
    ('schimbare_router', 'Schimbare Router'),
    ('schimbare_ip', 'Schimbare IP'),
    ('schimbare_hdd', 'Schimbare HDD'),
    ('schimbare_sursa', 'Schimbare Sursă'),
    ('dezinstalare', 'Dezinstalare'),
    ('suplimentare', 'Suplimentare'),
    ('repozitionare', 'Repozitionare'),
    ('restart', 'Restart'),
    ('Alta', 'Alta (Custom)')
]

# --- ALGORITMUL MAGIC BAZAT PE CULORI ---
def incarca_locatii():
    cale_folder = r'G:\My Drive\.biblia*' 
    fisiere = glob.glob(cale_folder)
    
    if not fisiere:
        return [('Implicit', 'Nu am găsit fișiere .biblia în Drive')]
    
    ultimul_fisier = max(fisiere, key=os.path.getmtime)
    
    try:
        wb = openpyxl.load_workbook(ultimul_fisier, data_only=True)
        ws = wb.active
        grupuri_culori = []
        grup_curent = []
        culoare_curenta = None
        
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            val = cell.value
            if not val: continue
            val_str = str(val).upper().lstrip('?').strip()
            culoare = str(cell.fill.start_color.index) if (cell.fill and cell.fill.start_color) else f"Alb_{row}"
            if culoare in ['00000000', '0']: culoare = f"Alb_{row}"
            
            if culoare == culoare_curenta:
                grup_curent.append(val_str)
            else:
                if grup_curent: grupuri_culori.append(grup_curent)
                grup_curent = [val_str]
                culoare_curenta = culoare
        if grup_curent: grupuri_culori.append(grup_curent)
            
        locatii_finale = set()
        fallback_pattern = r'\b(LIFT|SCARA|SC|BLOC|BL|CORP|C\d+|ETAJ|ET|AP|PARTER|SUBSOL|PARCARE|POARTA|BARIERA|ACCES|LATURA|STANGA|DREAPTA|MAGAZIE|SPATE|FATA|SMART\d*|CENTRU|FOTO|BUCATARIE|IP|TB\d*|INTERFON|EXT|INT|MIKROTIK|BARACA|SCULE|TGC|TARC|FIER|LATERAL|AUTO|PIETONAL|NVR|DVR|MIJLOC|CORE\d*|CAMERA|CAM|HOL|SERVER|RACK|GARAJ|RECEPTIE|BIROU|SALA|\d{1,2}MP|[A-Z])\b'
        
        for grup in grupuri_culori:
            if len(grup) == 1:
                parti = re.split(fallback_pattern, grup[0])
                titlu = re.sub(r'[.-]$', '', parti[0].strip()).strip()
            else:
                cuvinte_primul = grup[0].split()
                cuvinte_comune = []
                for i, cuv in enumerate(cuvinte_primul):
                    este_peste_tot = all(i < len(text_altul.split()) and text_altul.split()[i] == cuv for text_altul in grup[1:])
                    if este_peste_tot: cuvinte_comune.append(cuv)
                    else: break
                titlu = re.sub(r'[.-]$', '', " ".join(cuvinte_comune)).strip()
            if titlu and not titlu.replace('-', '').isdigit() and titlu != 'NAN':
                locatii_finale.add(titlu.title())
        return [(l, l) for l in sorted(list(locatii_finale))]
    except Exception as e:
        return [('Eroare', f'Eroare script culori: {str(e)}')]

# --- FORMULARE ---

class AdaugaUtilizatorForm(forms.ModelForm):
    parola = forms.CharField(widget=forms.PasswordInput)
    class Meta:
        model = User
        fields = ['username', 'email', 'role']

class TrimiteMesajForm(forms.ModelForm):
    class Meta:
        model = Mesaj
        fields = ['destinatar', 'continut']

class AdaugaTaskForm(forms.ModelForm):
    atribuit_catre = forms.ModelMultipleChoiceField(
        queryset=User.objects.filter(role__in=['tehnician', 'sofer', 'gestionar', 'inginer']),
        widget=forms.SelectMultiple(attrs={'class': 'form-select', 'size': '5'}),
        label="Angajați (ține CTRL pentru mai mulți)"
    )
    
    # 🔥 AM ADĂUGAT CÂMPUL SUPERVIZOR AICI 🔥
    supervizor = forms.ModelChoiceField(
        queryset=User.objects.filter(role__in=['supervizor', 'manager', 'superadmin']),
        widget=forms.Select(attrs={'class': 'form-select'}),
        label="Supervizor / Control (Opțional)",
        required=False
    )
    
    locatie = forms.CharField(
        widget=forms.Select(choices=incarca_locatii(), attrs={'class': 'form-select'}),
        label="Locație"
    )
    
    actiune_select = forms.MultipleChoiceField(
        choices=ACTIUNI_LISTA, 
        widget=forms.SelectMultiple(attrs={'class': 'form-select', 'size': '4'}),
        label="Acțiuni (ține CTRL pentru mai multe)",
        required=False
    )
    
    actiune_custom = forms.CharField(
        required=False, 
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Scrie aici dacă ai ales "Alta"'}),
        label="Acțiune Custom"
    )

    class Meta:
        model = Task
        # 🔥 AM ADĂUGAT 'supervizor' ÎN LISTA DE CÂMPURI 🔥
        fields = ['atribuit_catre', 'supervizor', 'locatie', 'deadline', 'descriere', 'observatii']
        widgets = {
            'deadline': forms.DateTimeInput(attrs={'type': 'datetime-local', 'class': 'form-control'}, format='%Y-%m-%dT%H:%M'),
            'descriere': forms.Textarea(attrs={'class': 'form-control', 'rows': 4, 'placeholder': 'Scrie aici detaliile...'}),
            'observatii': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Notează aici...'}),
        }

    # --- AICI ESTE FUNCȚIA NOUĂ CARE MEMOREAZĂ LOCAȚIILE SCRISE MANUAL ---
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # 1. Luăm locațiile vechi (din Excel/funcția ta)
        choices_actuale = list(self.fields['locatie'].widget.choices)
        nume_locatii_actuale = [str(c[0]) for c in choices_actuale]
        
        # 2. Căutăm în baza de date locațiile unice scrise manual în trecut
        locatii_db = Task.objects.exclude(locatie__isnull=True).exclude(locatie__exact='').values_list('locatie', flat=True).distinct()
        
        # 3. Le combinăm: dacă locația din baza de date nu e în Excel, o adăugăm în listă
        for loc in locatii_db:
            if str(loc) not in nume_locatii_actuale:
                choices_actuale.append((loc, loc))
                
        # 4. Punem lista finală, sortată alfabetic, înapoi în formular
        self.fields['locatie'].widget.choices = sorted(choices_actuale, key=lambda x: str(x[0]))
    # ----------------------------------------------------------------------

    def clean(self):
        cleaned_data = super().clean()
        actiuni = cleaned_data.get('actiune_select', [])
        if 'Alta' in actiuni:
            custom = cleaned_data.get('actiune_custom')
            if custom:
                actiuni.remove('Alta')
                actiuni.append(custom)
        cleaned_data['actiune'] = ", ".join(actiuni)
        return cleaned_data
        
    def save(self, commit=True):
        task = super().save(commit=False)
        task.actiune = self.cleaned_data.get('actiune', '')
        if commit:
            task.save()
            self.save_m2m()
        return task

class FinalizareTaskForm(forms.ModelForm):
    class Meta:
        model = Task
        fields = ['raport_finalizare']
        widgets = {
            'raport_finalizare': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        }

class ReminderForm(forms.ModelForm):
    class Meta:
        model = Reminder
        fields = ['titlu', 'detalii', 'data_reminder']
        widgets = {
            'titlu': forms.TextInput(attrs={'class': 'form-control'}),
            'detalii': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
            'data_reminder': forms.DateTimeInput(attrs={'type': 'datetime-local', 'class': 'form-control'}),
        }
    
class RaportSupervizorForm(forms.ModelForm):
    class Meta:
        model = RaportSupervizor
        fields = ['angajat_evaluat', 'stare_lucrare', 'ce_s_a_facut', 'ce_nu_s_a_facut']
        widgets = {
            'angajat_evaluat': forms.Select(attrs={'class': 'form-select'}),
            'stare_lucrare': forms.Select(attrs={'class': 'form-select'}),
            'ce_s_a_facut': forms.Textarea(attrs={'class': 'form-control', 'rows': 4, 'placeholder': 'Descrie ce s-a lucrat...'}),
            'ce_nu_s_a_facut': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Descrie problemele sau ce a rămas nefăcut...'}),
        }